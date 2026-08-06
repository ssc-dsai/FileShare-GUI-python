# Classification/export_results.py
"""
Write classification results to CSV + colour-coded Excel.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def save_results(df: pd.DataFrame, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_path = output_dir / "classification_results.csv"
    excel_path = output_dir / "classification_results.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV saved: {csv_path}")

    df.to_excel(excel_path, index=False, engine="openpyxl")
    _format_excel(excel_path, df)
    logger.info(f"Excel saved: {excel_path}")

    return csv_path, excel_path


def _format_excel(excel_path: Path, df: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="C6EFCE")
    med_fill = PatternFill("solid", fgColor="FFEB9C")
    low_fill = PatternFill("solid", fgColor="FFC7CE")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    # Colour confidence_category column if present
    try:
        conf_col = list(df.columns).index("confidence_category") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, conf_col)
            val = str(cell.value or "").lower()
            if val == "high":
                cell.fill = high_fill
            elif val == "medium":
                cell.fill = med_fill
            elif val == "low":
                cell.fill = low_fill
    except ValueError:
        pass

    # Reasonable column widths for excerpt columns
    for idx, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(idx)
        if "Excerpt" in col_name:
            ws.column_dimensions[letter].width = 50
        elif col_name in ("original_path", "Records"):
            ws.column_dimensions[letter].width = 40
        else:
            ws.column_dimensions[letter].width = min(28, max(12, len(col_name) + 2))

    wb.save(excel_path)