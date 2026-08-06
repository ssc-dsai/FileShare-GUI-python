# DeDuplication/0_dedup_analysis.py
# Launched from Gradio via backend.runners.run_dedup_analysis()
# Or manually: python DeDuplication/0_dedup_analysis.py

import hashlib
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# ---------------------------------------------------------------------------
# Project root + central config
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_config import (
    SOURCE_DOCS_DIR as SOURCE_DIR,
    DEDUPS_DIR,
    TRIVIAL_SUBJECTS as TRIVIAL_SUBJECTS_FILE,
    EMBEDDING_MODEL_PATH,
)

# Optional extractors
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    from docx import Document
except ImportError:
    Document = None

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SIMILARITY_THRESHOLD = 0.95

DEDUPS_DIR.mkdir(parents=True, exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_EXCEL = DEDUPS_DIR / f"deduplication_review_{timestamp}.xlsx"
LOG_FILE = DEDUPS_DIR / f"dedup_analysis_{timestamp}.log"

# ---------------------------------------------------------------------------
# Logging (captured by Gradio)
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

logger.info("=== Deduplication Analysis Started ===")
logger.info(f"Source directory : {SOURCE_DIR}")
logger.info(f"Output folder    : {DEDUPS_DIR}")
logger.info(f"Trivial subjects : {TRIVIAL_SUBJECTS_FILE}")

# Load trivial subjects
if TRIVIAL_SUBJECTS_FILE.exists():
    trivial_subjects = [
        line.strip()
        for line in TRIVIAL_SUBJECTS_FILE.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    logger.info(f"Loaded {len(trivial_subjects)} trivial subjects")
else:
    trivial_subjects = []
    logger.warning("trivial_subjects.txt not found – trivial detection disabled")

# Load embedding model from local disk (fully offline)
logger.info(f"Loading embedding model from local path: {EMBEDDING_MODEL_PATH}")

if not EMBEDDING_MODEL_PATH.exists():
    raise FileNotFoundError(
        f"Local embedding model not found at:\n{EMBEDDING_MODEL_PATH}\n"
        "Please download it first."
    )

embedder = SentenceTransformer(str(EMBEDDING_MODEL_PATH))
logger.info("Embedding model ready (local, offline)")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def extract_text_from_file(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    try:
        if suffix == ".txt":
            return file_path.read_text(encoding="utf-8", errors="ignore")
        elif suffix == ".pdf" and fitz:
            doc = fitz.open(file_path)
            text = "\n".join(page.get_text("text") for page in doc)
            doc.close()
            return text
        elif suffix in (".docx", ".doc") and Document:
            if suffix == ".doc":
                return file_path.name  # old binary .doc not supported
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)
        else:
            return file_path.name
    except Exception as e:
        logger.warning(f"Text extraction failed for {file_path.name}: {e}")
        return file_path.name


def extract_hyperlinks(text: str) -> str:
    pattern = r"(https?://\S+|www\.\S+|\b[a-zA-Z0-9-]+\.(com|ca|org|net|gov|edu)\b)"
    links = re.findall(pattern, text, re.IGNORECASE)
    cleaned = {link[0] if isinstance(link, tuple) else link for link in links}
    return ", ".join(sorted(cleaned)) if cleaned else ""


def is_trivial_content(text: str) -> bool:
    """Pure-Python trivial detector (keyword match against trivial_subjects.txt)."""
    if not trivial_subjects or not text.strip():
        return False
    sample = text[:3000].lower()
    for subject in trivial_subjects:
        key = subject.strip().lower()
        if len(key) >= 4 and key in sample:
            return True
    return False


def compute_file_hash(file_path: Path) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------
logger.info(f"Scanning source directory: {SOURCE_DIR}")
files = [p for p in SOURCE_DIR.rglob("*.*") if p.is_file()]
logger.info(f"Found {len(files)} files to analyze")

records = []
embeddings = []
file_paths = []

for file_path in files:
    try:
        stat = file_path.stat()
        hash_value = compute_file_hash(file_path)
        text = extract_text_from_file(file_path)
        is_trivial = is_trivial_content(text)

        emb = embedder.encode(text[:3000], normalize_embeddings=True)

        records.append(
            {
                "filename": file_path.name,
                "original_path": str(file_path),
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "creation_time": datetime.fromtimestamp(stat.st_ctime).strftime(
                    "%Y-%m-%d %H:%M"
                ),
                "last_modified": datetime.fromtimestamp(stat.st_mtime).strftime(
                    "%Y-%m-%d %H:%M"
                ),
                "Is_Trivial": "Yes" if is_trivial else "No",
                "Hash_Value": hash_value,
            }
        )
        embeddings.append(emb)
        file_paths.append(file_path)
    except Exception as e:
        logger.warning(f"Failed to process {file_path.name}: {e}")

if not embeddings:
    logger.error("No files processed.")
    print("No files found to analyze.")
    sys.exit(0)

df = pd.DataFrame(records)
embeddings = np.array(embeddings)

logger.info("Computing similarity matrix...")
sim_matrix = cosine_similarity(embeddings)

# -------------------- Clustering --------------------
visited = set()
cluster_id = 0
df["Cluster_ID"] = "Unique"
df["Is_Master"] = False
df["Similarity_Score"] = 1.0
df["Similarity_Found"] = ""
df["Discrepancy"] = ""

for i in range(len(df)):
    if i in visited:
        continue

    cluster = [i]
    visited.add(i)

    for j in range(i + 1, len(df)):
        if sim_matrix[i, j] >= SIMILARITY_THRESHOLD and j not in visited:
            cluster.append(j)
            visited.add(j)

    if len(cluster) > 1:
        cluster_id += 1
        cid = f"DUP-{cluster_id:04d}"

        cluster_files = [(idx, file_paths[idx]) for idx in cluster]
        # Oldest file becomes master
        master_idx = min(
            cluster_files,
            key=lambda x: (Path(x[1]).stat().st_ctime, Path(x[1]).stat().st_mtime),
        )[0]

        for idx in cluster:
            df.loc[idx, "Cluster_ID"] = cid
            df.loc[idx, "Is_Master"] = idx == master_idx
            score = float(sim_matrix[master_idx, idx])
            df.loc[idx, "Similarity_Score"] = round(score, 4)

            if (
                abs(score - 1.0) < 0.0001
                and df.loc[idx, "Hash_Value"] == df.loc[master_idx, "Hash_Value"]
            ):
                df.loc[idx, "Similarity_Found"] = "Exact duplicate (hash + content)"
            else:
                df.loc[idx, "Similarity_Found"] = f"{score:.2%} semantic similarity"

            master_path = Path(df.loc[master_idx, "original_path"])
            current_path = Path(df.loc[idx, "original_path"])
            if df.loc[idx, "filename"] != df.loc[master_idx, "filename"]:
                df.loc[idx, "Discrepancy"] = "Different filename"
            elif master_path.stat().st_mtime != current_path.stat().st_mtime:
                df.loc[idx, "Discrepancy"] = "Different modification date"
            else:
                df.loc[idx, "Discrepancy"] = "Minor content variations"
    else:
        df.loc[i, "Similarity_Found"] = "Unique file"
        df.loc[i, "Discrepancy"] = "No similar documents found"
        df.loc[i, "Is_Master"] = True


def get_recommended_action(row):
    if row["Is_Trivial"] == "Yes":
        return "Review (trivial content)"
    if row["Cluster_ID"] == "Unique" or row["Is_Master"]:
        return "Keep as Master"
    return "Delete"


df["Recommended_Action"] = df.apply(get_recommended_action, axis=1)
df["User_Confirmed_Delete"] = df.apply(
    lambda row: "Yes" if row["Recommended_Action"] == "Delete" else "", axis=1
)

df["Exact_Duplicate"] = df.duplicated(subset=["Hash_Value"], keep=False)
df["Near_Duplicate"] = (df["Similarity_Score"] >= SIMILARITY_THRESHOLD) & (
    df["Cluster_ID"] != "Unique"
)
df["Hash_Type"] = "SHA-256"

df["downstream_hyperlinks"] = df["original_path"].apply(
    lambda p: extract_hyperlinks(extract_text_from_file(Path(p)))
)

# Final column order
cols = [
    "Cluster_ID",
    "Is_Master",
    "filename",
    "original_path",
    "size_mb",
    "creation_time",
    "last_modified",
    "Exact_Duplicate",
    "Near_Duplicate",
    "Similarity_Score",
    "Similarity_Found",
    "Discrepancy",
    "Recommended_Action",
    "User_Confirmed_Delete",
    "Is_Trivial",
    "downstream_hyperlinks",
    "Hash_Value",
    "Hash_Type",
]
df = df[cols]

# ---------------------------------------------------------------------------
# Write Excel
# ---------------------------------------------------------------------------
with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Duplicate_Clusters", index=False)

    ws = writer.sheets["Duplicate_Clusters"]
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row in range(2, len(df) + 2):
        # filename hyperlink (column C = 3)
        cell = ws.cell(row=row, column=3)
        cell.hyperlink = df.iloc[row - 2]["original_path"]
        cell.value = df.iloc[row - 2]["filename"]
        cell.style = "Hyperlink"

        # colour Recommended_Action (column M = 13)
        action_cell = ws.cell(row=row, column=13)
        if action_cell.value in ("Delete", "Review (trivial content)"):
            action_cell.fill = red_fill

    # Summary sheet
    summary = pd.DataFrame(
        {
            "Metric": [
                "Total Files Scanned",
                "Duplicate Clusters Found",
                "Files Recommended for Deletion",
                "Trivial Content Flagged",
                "Unique Files",
                "Run Timestamp",
            ],
            "Value": [
                len(df),
                df["Cluster_ID"].nunique()
                - (1 if "Unique" in df["Cluster_ID"].values else 0),
                (df["Recommended_Action"] == "Delete").sum(),
                (df["Is_Trivial"] == "Yes").sum(),
                (df["Cluster_ID"] == "Unique").sum(),
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ],
        }
    )
    summary.to_excel(writer, sheet_name="Summary", index=False)

logger.info(f"Analysis complete. Excel report saved to: {OUTPUT_EXCEL}")
print("\n✅ Deduplication analysis complete!")
print(f"   Excel:  {OUTPUT_EXCEL}")
print(f"   Log:    {LOG_FILE}")
print(f"   Source: {SOURCE_DIR}")